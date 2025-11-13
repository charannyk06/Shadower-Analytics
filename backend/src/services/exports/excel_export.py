"""Excel export functionality."""

from typing import List, Dict, Optional
from datetime import datetime, date
import io


def export_to_excel(
    data: List[Dict],
    output_path: Optional[str] = None,
    sheet_name: str = "Sheet1",
    columns: Optional[List[str]] = None
) -> Optional[str]:
    """Export data to Excel format.

    Args:
        data: List of dictionaries to export
        output_path: File path to write to
        sheet_name: Name of the worksheet
        columns: Optional list of column names to include

    Returns:
        File path if output_path is provided, None otherwise

    Note:
        This is a placeholder. In production, you would need to install:
        - openpyxl for .xlsx files
        - xlsxwriter for enhanced features
    """
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel export. "
            "Install it with: pip install openpyxl"
        )

    if not data:
        return None

    # Determine columns
    if columns is None:
        columns = list(data[0].keys())

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Style header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Write headers
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font

    # Write data
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            value = row_data.get(col_name, "")

            # Convert datetime objects to ISO format
            if isinstance(value, (datetime, date)):
                value = value.isoformat()

            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to file or return bytes
    if output_path:
        wb.save(output_path)
        return output_path
    else:
        # Return as bytes in memory
        virtual_workbook = io.BytesIO()
        wb.save(virtual_workbook)
        virtual_workbook.seek(0)
        return virtual_workbook.getvalue()


def export_to_excel_multi_sheet(
    data_dict: Dict[str, List[Dict]],
    output_path: str
) -> str:
    """Export multiple datasets to Excel with multiple sheets.

    Args:
        data_dict: Dictionary mapping sheet names to data lists
        output_path: File path to write to

    Returns:
        File path
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel export. "
            "Install it with: pip install openpyxl"
        )

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for sheet_name, data in data_dict.items():
        if not data:
            continue

        ws = wb.create_sheet(title=sheet_name)
        columns = list(data[0].keys())

        # Write headers
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font

        # Write data
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                value = row_data.get(col_name, "")
                if isinstance(value, (datetime, date)):
                    value = value.isoformat()
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    return output_path


def stream_to_excel(
    data_generator,
    output_path: str,
    columns: List[str],
    sheet_name: str = "Sheet1"
) -> int:
    """Stream data to Excel file.

    Args:
        data_generator: Generator that yields dictionaries
        output_path: File path to write to
        columns: Column names
        sheet_name: Name of the worksheet

    Returns:
        Number of rows written
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel export. "
            "Install it with: pip install openpyxl"
        )

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Write headers
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font

    # Write data
    row_count = 0
    for row_idx, row_data in enumerate(data_generator, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            value = row_data.get(col_name, "")
            if isinstance(value, (datetime, date)):
                value = value.isoformat()
            ws.cell(row=row_idx, column=col_idx, value=value)
        row_count += 1

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    return row_count
